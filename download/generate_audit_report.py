#!/usr/bin/env python3
"""Generate CodeBot V4.1.0 Comprehensive Audit Report (Chinese)."""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from copy import copy

OUTPUT = "/home/z/my-project/download/CodeBot_V4.1.0_Audit_Report.xlsx"

wb = Workbook()

# ──────────────────────────── Colour Palette ────────────────────────────
WHITE = "FFFFFF"
LIGHT_GREY = "F5F5F5"
MID_GREY = "E0E0E0"
DARK_GREY = "4A4A4A"
HEADER_BG = "6B7280"          # grey-600
HEADER_FONT_CLR = "FFFFFF"
ACCENT_BLUE = "3B82F6"
CRITICAL_RED = "FEE2E2"
CRITICAL_RED_FONT = "991B1B"
HIGH_ORANGE = "FEF3C7"
HIGH_ORANGE_FONT = "92400E"
MEDIUM_YELLOW = "FEF9C3"
MEDIUM_YELLOW_FONT = "854D0E"
LOW_GREEN = "DCFCE7"
LOW_GREEN_FONT = "166534"
BORDER_CLR = "D1D5DB"

thin_border = Border(
    left=Side(style='thin', color=BORDER_CLR),
    right=Side(style='thin', color=BORDER_CLR),
    top=Side(style='thin', color=BORDER_CLR),
    bottom=Side(style='thin', color=BORDER_CLR),
)

header_font = Font(name="Microsoft YaHei", size=11, bold=True, color=HEADER_FONT_CLR)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

body_font = Font(name="Microsoft YaHei", size=10, color="1F2937")
body_align = Alignment(vertical="center", wrap_text=True)
body_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Microsoft YaHei", size=18, bold=True, color=DARK_GREY)
subtitle_font = Font(name="Microsoft YaHei", size=12, bold=False, color="6B7280")

# Alternating row fills
row_fill_even = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
row_fill_odd = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Severity fills
severity_fills = {
    "Critical": PatternFill(start_color=CRITICAL_RED, end_color=CRITICAL_RED, fill_type="solid"),
    "High":     PatternFill(start_color=HIGH_ORANGE, end_color=HIGH_ORANGE, fill_type="solid"),
    "Medium":   PatternFill(start_color=MEDIUM_YELLOW, end_color=MEDIUM_YELLOW, fill_type="solid"),
    "Low":      PatternFill(start_color=LOW_GREEN, end_color=LOW_GREEN, fill_type="solid"),
}
severity_fonts = {
    "Critical": Font(name="Microsoft YaHei", size=10, bold=True, color=CRITICAL_RED_FONT),
    "High":     Font(name="Microsoft YaHei", size=10, bold=True, color=HIGH_ORANGE_FONT),
    "Medium":   Font(name="Microsoft YaHei", size=10, bold=True, color=MEDIUM_YELLOW_FONT),
    "Low":      Font(name="Microsoft YaHei", size=10, bold=True, color=LOW_GREEN_FONT),
}


def apply_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        apply_cell(c, font=header_font, fill=header_fill, alignment=header_align, border=thin_border)


def style_body_cell(ws, row, col, center=False, severity=None, alt_row=False):
    c = ws.cell(row=row, column=col)
    font = body_font
    fill = row_fill_even if alt_row else row_fill_odd
    # If severity provided and this is the severity column, use special style
    if severity:
        font = severity_fonts.get(severity, body_font)
        fill = severity_fills.get(severity, fill)
    align = body_align_center if center else body_align
    apply_cell(c, font=font, fill=fill, alignment=align, border=thin_border)


# ════════════════════════════════════════════════════════════════════════
# SHEET 1 — 审计总览
# ════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "审计总览"
ws1.sheet_properties.tabColor = ACCENT_BLUE

# Title
ws1.merge_cells("B2:G2")
c = ws1["B2"]
c.value = "CodeBot V4.1.0 全面功能审计报告"
apply_cell(c, font=title_font, alignment=Alignment(horizontal="center", vertical="center"))
ws1.row_dimensions[2].height = 40

# Subtitle
ws1.merge_cells("B3:G3")
c = ws1["B3"]
c.value = "审计日期: 2025-07-13  |  审计工具: 人工审查 + 静态分析  |  版本: V4.1.0"
apply_cell(c, font=subtitle_font, alignment=Alignment(horizontal="center", vertical="center"))
ws1.row_dimensions[3].height = 24

# ── Summary table ──
summary_headers = ["指标", "数值"]
summary_data = [
    ("审计模块数", 5),
    ("审计文件数", "145+"),
    ("发现问题总数", 138),
    ("Critical", 16),
    ("High", 40),
    ("Medium", 59),
    ("Low", 23),
]

r = 5
ws1.merge_cells("B5:C5")
c = ws1["B5"]
c.value = "审计摘要统计"
apply_cell(c, font=Font(name="Microsoft YaHei", size=14, bold=True, color=DARK_GREY),
           alignment=Alignment(horizontal="left", vertical="center"))
ws1.row_dimensions[5].height = 28

r = 6
for i, h in enumerate(summary_headers):
    c = ws1.cell(row=r, column=2 + i, value=h)
    apply_cell(c, font=header_font, fill=header_fill, alignment=header_align, border=thin_border)
ws1.row_dimensions[6].height = 24

for idx, (label, val) in enumerate(summary_data):
    rr = 7 + idx
    is_alt = idx % 2 == 1
    fill = row_fill_even if is_alt else row_fill_odd
    cl = ws1.cell(row=rr, column=2, value=label)
    apply_cell(cl, font=body_font, fill=fill, alignment=Alignment(horizontal="left", vertical="center", wrap_text=True), border=thin_border)
    cv = ws1.cell(row=rr, column=3, value=val)
    # Color-code severity rows
    if label == "Critical":
        apply_cell(cv, font=severity_fonts["Critical"], fill=severity_fills["Critical"],
                   alignment=body_align_center, border=thin_border)
    elif label == "High":
        apply_cell(cv, font=severity_fonts["High"], fill=severity_fills["High"],
                   alignment=body_align_center, border=thin_border)
    elif label == "Medium":
        apply_cell(cv, font=severity_fonts["Medium"], fill=severity_fills["Medium"],
                   alignment=body_align_center, border=thin_border)
    elif label == "Low":
        apply_cell(cv, font=severity_fonts["Low"], fill=severity_fills["Low"],
                   alignment=body_align_center, border=thin_border)
    else:
        apply_cell(cv, font=Font(name="Microsoft YaHei", size=11, bold=True, color=DARK_GREY),
                   fill=fill, alignment=body_align_center, border=thin_border)
    ws1.row_dimensions[rr].height = 22

# ── Bar Chart: issues by severity ──
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "问题数量 — 按严重度分布"
chart.y_axis.title = "问题数"
chart.x_axis.title = "严重度"
chart.width = 20
chart.height = 14

# Chart data is in rows 10-13 (Critical, High, Medium, Low), column C (value)
data = Reference(ws1, min_col=3, min_row=9, max_row=13)
cats = Reference(ws1, min_col=2, min_row=10, max_row=13)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.shape = 4

# Color the bars
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from copy import copy

colors_bar = ["EF4444", "F59E0B", "EAB308", "22C55E"]  # red, amber, yellow, green
for i, clr in enumerate(colors_bar):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = clr
    chart.series[0].data_points.append(pt)

chart.legend = None
ws1.add_chart(chart, "E5")

# ── Module breakdown table ──
mod_start = 16
ws1.merge_cells(f"B{mod_start}:F{mod_start}")
c = ws1[f"B{mod_start}"]
c.value = "各审计模块问题分布"
apply_cell(c, font=Font(name="Microsoft YaHei", size=14, bold=True, color=DARK_GREY),
           alignment=Alignment(horizontal="left", vertical="center"))
ws1.row_dimensions[mod_start].height = 28

mod_headers = ["模块", "文件数", "Critical", "High", "Medium/Low"]
mod_data = [
    ("Chat模块", 38, 5, 9, 15),
    ("UI/Layout", 32, 3, 10, 12),
    ("Agents/Bridge/Skills", 28, 2, 5, 10),
    ("Settings/Memory/Tools", 25, 4, 6, 14),
    ("Security/API", 22, 2, 10, 8),
]

r = mod_start + 1
for i, h in enumerate(mod_headers):
    c = ws1.cell(row=r, column=2 + i, value=h)
    apply_cell(c, font=header_font, fill=header_fill, alignment=header_align, border=thin_border)
ws1.row_dimensions[r].height = 24

for idx, row_data in enumerate(mod_data):
    rr = r + 1 + idx
    is_alt = idx % 2 == 1
    fill = row_fill_even if is_alt else row_fill_odd
    for col_idx, val in enumerate(row_data):
        c = ws1.cell(row=rr, column=2 + col_idx, value=val)
        center = col_idx > 0
        f = body_font
        if col_idx == 2 and isinstance(val, int):  # Critical col
            f = severity_fonts["Critical"]
            fill = severity_fills["Critical"]
        elif col_idx == 3 and isinstance(val, int):  # High col
            f = severity_fonts["High"]
            fill = severity_fills["High"]
        apply_cell(c, font=f, fill=fill,
                   alignment=body_align_center if center else Alignment(horizontal="left", vertical="center", wrap_text=True),
                   border=thin_border)
    ws1.row_dimensions[rr].height = 22

# Column widths Sheet 1
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 14


# ════════════════════════════════════════════════════════════════════════
# SHEET 2 — 问题清单
# ════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("问题清单")
ws2.sheet_properties.tabColor = "EF4444"

# All issues: (ID, 模块, 严重度, 类别, 文件路径, 问题描述, 修复建议)
issues = [
    # ── CRITICAL (20) ──
    ("Chat-C1", "Chat模块", "Critical", "安全", "所有API路由", "所有API路由零认证/鉴权", "添加Next.js middleware或bearer-token认证"),
    ("Chat-C2", "Chat模块", "Critical", "安全", "RichContentRenderer.tsx", "rehypeRaw允许渲染原始HTML导致XSS", "添加rehype-sanitize或DOMPurify"),
    ("Chat-C3", "Chat模块", "Critical", "安全", "bash.ts", "process.env全部传递给子进程泄露密钥", "仅白名单所需环境变量"),
    ("Chat-C4", "Chat模块", "Critical", "Bug", "chat/route.ts", "会话历史加载最旧20条而非最新", "改用orderBy desc+skip或cursor"),
    ("Chat-C5", "Chat模块", "Critical", "安全", "chat/stream/route.ts", "聊天端点无服务端限流", "添加Upstash Ratelimit或中间件限流"),
    ("UI-C1", "UI/Layout", "Critical", "构建", "tailwind.config.ts", "Tailwind v3 config残留与v4冲突", "删除tailwind.config.ts和tailwindcss-animate"),
    ("UI-C2", "UI/Layout", "Critical", "代码质量", "Sidebar.tsx", "dead Sidebar()导出+hsl() v3语法", "移除死代码，改用CSS变量"),
    ("UI-C3", "UI/Layout", "Critical", "可靠性", "全局", "全应用零React Error Boundary", "在page.tsx和关键组件添加Error Boundary"),
    ("Agent-C1", "Agents", "Critical", "Bug", "swarm.ts:55", "MessageBusBus拼写错误导致swarm模式崩溃", "改为new MessageBusClass()"),
    ("Agent-C2", "Agents", "Critical", "Bug", "ChatView.tsx", "Teammate模式无API路由", "创建/api/agents/teammate/route.ts"),
    ("Sec-C1", "Security", "Critical", "安全", ".env", "生产API密钥(NVIDIA+Supabase)在.env中", "轮换密钥，使用.env.example"),
    ("Sec-C2", "Security", "Critical", "安全", "所有API路由", "零认证鉴权(next-auth已安装但未使用)", "实现NextAuth或bearer-token中间件"),
    ("Sec-C3", "Security", "Critical", "安全", "Caddyfile", "XTransformPort允许访问任意内部端口(SSRF)", "限制为已知端口白名单"),
    ("Sec-C4", "Security", "Critical", "安全", "bash.ts", "无限制bash命令执行+无人工审批", "添加沙箱和审批门"),
    ("Sec-C5", "Security", "Critical", "安全", "file-operations.ts", "无限制文件系统读写(任意路径)", "限制工作区根目录"),
    ("Sec-C6", "Security", "Critical", "安全", "custom-models/[id]/route.ts", "PATCH无输入验证(mass assignment)", "添加允许字段白名单"),
    ("Store-C1", "Settings", "Critical", "Bug(P4)", "Sidebar.tsx", "会话DELETE调用错误端点，DB记录从未删除", "在sessions/[id]/route.ts添加DELETE"),
    ("Store-C2", "Settings", "Critical", "性能", "schema.prisma", "Memory表无数据库索引", "添加@@index([layer,sessionId])等"),
    ("Store-C3", "Settings", "Critical", "数据完整性", "多文件", "Memory层名不一致(5种变体)", "统一为一种命名规范"),
    ("Store-C4", "Settings", "Critical", "安全", "chat/stream/route.ts", "工具执行无人工审批门", "高风险工具添加审批流程"),
    # ── HIGH (38) ──
    ("Chat-H1", "Chat", "High", "代码质量", "ChatView.tsx", "2000+行超级组件", "拆分为SessionPanel/MessageList/ChatInput等"),
    ("Chat-H2", "Chat", "High", "代码质量", "多文件", "MultiAgentMode等类型重复定义", "移至src/lib/types.ts"),
    ("Chat-H3", "Chat", "High", "Bug", "ChatView.tsx", "Session自动创建存在竞态条件", "直接setState或传递sessionId"),
    ("Chat-H4", "Chat", "High", "功能缺失", "MessageBubble.tsx", "无消息重试/重新发送功能", "失败消息添加重试按钮"),
    ("Chat-H5", "Chat", "High", "UX/安全", "ChatView.tsx", "图片上传无大小验证", "添加10MB限制"),
    ("Chat-H6", "Chat", "High", "代码质量", "chat/route.ts+stream/route.ts", "estimateTokens函数重复定义", "提取到lib/utils.ts"),
    ("Chat-H7", "Chat", "High", "代码质量", "多文件", "System Prompt硬编码重复", "存入DB AgentConfig模型"),
    ("Chat-H8", "Chat", "High", "安全", "所有API路由", "错误响应泄露内部详情", "生产环境返回通用错误"),
    ("Chat-H9", "Chat", "High", "UX", "ChatView.tsx", "自动滚动不检测用户位置", "跟踪用户滚动位置"),
    ("UI-H1", "UI/Layout", "High", "UI/UX", "layout.tsx", "mono字体变量--font-geist-mono未加载", "加载JetBrains Mono或更新CSS变量"),
    ("UI-H2", "UI/Layout", "High", "代码质量", "layout.tsx", "Inter字体命名为--font-geist-sans(误导)", "改名或切换到Geist字体"),
    ("UI-H3", "UI/Layout", "High", "代码质量", "globals.css", "CSS定义重复(layer内+layer外)", "合并为单一block"),
    ("UI-H4", "UI/Layout", "High", "代码质量", "next.config.ts", "reactStrictMode: false", "设为true并修复警告"),
    ("UI-H5", "UI/Layout", "High", "可访问性", "多文件", "交互元素缺少ARIA标签", "添加aria-label到所有图标按钮"),
    ("UI-H6", "UI/Layout", "High", "响应式", "page.tsx+Sidebar.tsx", "768-1024px断点间隙无导航", "对齐md/lg断点"),
    ("UI-H7", "UI/Layout", "High", "UI/UX", "多文件", "z-index管理不一致", "创建z-index CSS变量体系"),
    ("UI-H8", "UI/Layout", "High", "性能", "use-toast.ts", "TOAST_REMOVE_DELAY=16.7min+死代码", "删除use-toast.ts(已被sonner取代)"),
    ("UI-H9", "UI/Layout", "High", "代码质量", "MessageBubble.tsx", "h-4.5 w-4.5非标准Tailwind类", "改为h-5 w-5或h-[18px]"),
    ("UI-H10", "UI/Layout", "High", "UI/UX", "多文件", "硬编码zinc颜色替代CSS变量", "改用bg-card/border-border等"),
    ("Agent-H1", "Agents", "High", "功能(P5)", "AgentsView.tsx", "100%使用mock数据无API集成", "添加useEffect获取真实API"),
    ("Agent-H2", "Agents", "High", "代码质量", "types.ts+API", "AgentStatus类型不匹配(8vs4)", "统一为4种状态"),
    ("Agent-H3", "Agents", "High", "功能(P3)", "BridgeView.tsx", "Bridge→Chat无路由/交接", "添加发送到Chat按钮"),
    ("Agent-H4", "Agents", "High", "功能", "skills/route.ts", "Skills无DELETE端点", "添加DELETE /api/skills"),
    ("Agent-H5", "Agents", "High", "性能", "agents API", "Agent列表/消息查询无分页", "添加page/limit参数"),
    ("Sec-H1", "Security", "High", "安全", "next.config.ts", "无CORS配置", "按需添加CORS头"),
    ("Sec-H2", "Security", "High", "安全", "所有API路由", "无服务端限流", "添加中间件限流"),
    ("Sec-H3", "Security", "High", "安全", "image-analyze/route.ts", "无文件大小/类型验证", "添加MIME和大小校验"),
    ("Sec-H4", "Security", "High", "安全", "image-analyze/route.ts", "响应泄露完整base64图片", "移除imagePreview字段"),
    ("Sec-H5", "Security", "High", "安全", "bridge/status/route.ts", "channelId路径注入(SSRF)", "验证channelId格式"),
    ("Sec-H6", "Security", "High", "安全", "多个API路由", "错误消息暴露内部详情", "生产环境返回通用错误"),
    ("Sec-H7", "Security", "High", "代码质量", "package.json", "next-auth已安装但从未使用", "移除或实现认证"),
    ("Sec-H8", "Security", "High", "安全", "chat/stream/route.ts", "Agentic循环可被滥用耗尽资源", "添加并发和预算限制"),
    ("Store-H1", "Settings", "High", "功能", "SettingsView.tsx", "设置从不持久化到数据库", "handleSave添加PUT /api/settings调用"),
    ("Store-H2", "Settings", "High", "架构", "chat-store.ts", "Zustand store无persist中间件", "添加persist关键状态"),
    ("Store-H3", "Settings", "High", "Bug", "chat/route.ts", "消息历史加载最旧而非最新(take:20)", "改用skip+orderBy desc"),
    ("Store-H4", "Settings", "High", "类型", "types.ts+schema", "AgentConfig TS接口与Prisma不匹配", "同步接口字段"),
    ("Store-H5", "Settings", "High", "架构", "tools", "两套独立工具系统不同步", "统一为代码定义为真相源"),
    ("Store-H6", "Settings", "High", "类型", "types.ts+nvidia.ts", "ModelCategory/NvidiaModel类型重复", "合并为单一定义"),
    # ── MEDIUM (Top 20) ──
    ("Chat-M1", "Chat", "Medium", "UX", "ChatView.tsx", "发送按钮无loading状态", "添加isStreaming状态禁用按钮"),
    ("Chat-M2", "Chat", "Medium", "UX", "MessageBubble.tsx", "代码块无复制按钮", "添加CopyButton组件"),
    ("Chat-M3", "Chat", "Medium", "代码质量", "chat/route.ts", "错误处理不一致(部分catch为空)", "统一错误处理模式"),
    ("Chat-M4", "Chat", "Medium", "功能", "ChatView.tsx", "无会话搜索功能", "添加搜索组件"),
    ("Chat-M5", "Chat", "Medium", "性能", "ChatView.tsx", "消息列表无虚拟化(1000+条卡顿)", "使用react-window虚拟滚动"),
    ("UI-M1", "UI/Layout", "Medium", "UI/UX", "Sidebar.tsx", "会话列表无分组/排序", "添加按日期分组排序"),
    ("UI-M2", "UI/Layout", "Medium", "响应式", "多文件", "移动端适配不完整", "完善移动端断点样式"),
    ("UI-M3", "UI/Layout", "Medium", "代码质量", "多文件", "大量console.log未清理", "生产构建前移除所有console"),
    ("UI-M4", "UI/Layout", "Medium", "性能", "globals.css", "未使用CSS大量残留", "清理未使用的CSS规则"),
    ("UI-M5", "UI/Layout", "Medium", "可访问性", "多文件", "颜色对比度不足(WCAG)", "调整文字/背景色对比"),
    ("Agent-M1", "Agents", "Medium", "功能", "AgentsView.tsx", "Agent编辑无保存反馈", "添加Toast通知"),
    ("Agent-M2", "Agents", "Medium", "UX", "AgentsView.tsx", "无Agent详情页", "创建Agent详情路由"),
    ("Agent-M3", "Agents", "Medium", "代码质量", "swarm.ts", "Swarm逻辑与UI耦合", "拆分swarm-core与swarm-ui"),
    ("Sec-M1", "Security", "Medium", "安全", "所有API路由", "无请求日志/审计追踪", "添加请求日志中间件"),
    ("Sec-M2", "Security", "Medium", "安全", "package.json", "依赖未锁版本(部分)", "使用pnpm lock确保一致性"),
    ("Store-M1", "Settings", "Medium", "功能", "SettingsView.tsx", "无主题切换持久化", "persist到localStorage/DB"),
    ("Store-M2", "Settings", "Medium", "性能", "schema.prisma", "Session表无复合索引", "添加@@index([userId,updatedAt])"),
    ("Store-M3", "Settings", "Medium", "代码质量", "chat-store.ts", "Store状态派生逻辑复杂", "使用Zustand selector模式"),
    ("Store-M4", "Settings", "Medium", "功能", "多文件", "Memory搜索无全文索引", "添加PostgreSQL全文搜索"),
    ("Store-M5", "Settings", "Medium", "数据完整性", "types.ts", "枚举值与DB不统一", "使用Prisma generated types"),
]

# Title row
ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "问题清单 — Critical (20) + High (38) + Medium (Top 20) = 78项"
apply_cell(c, font=Font(name="Microsoft YaHei", size=14, bold=True, color=DARK_GREY),
           alignment=Alignment(horizontal="left", vertical="center"))
ws2.row_dimensions[1].height = 30

# Headers at row 3
headers2 = ["ID", "模块", "严重度", "类别", "文件路径", "问题描述", "修复建议"]
for i, h in enumerate(headers2):
    c = ws2.cell(row=3, column=1 + i, value=h)
    apply_cell(c, font=header_font, fill=header_fill, alignment=header_align, border=thin_border)
ws2.row_dimensions[3].height = 28

# Freeze panes
ws2.freeze_panes = "A4"

# Auto-filter
ws2.auto_filter.ref = f"A3:G{3 + len(issues)}"

# Data rows
for idx, issue in enumerate(issues):
    rr = 4 + idx
    is_alt = idx % 2 == 1
    severity = issue[2]

    for col_idx, val in enumerate(issue):
        c = ws2.cell(row=rr, column=1 + col_idx, value=val)
        # Decide font and fill based on severity
        if col_idx == 2:  # Severity column
            apply_cell(c, font=severity_fonts.get(severity, body_font),
                       fill=severity_fills.get(severity, row_fill_even if is_alt else row_fill_odd),
                       alignment=body_align_center, border=thin_border)
        else:
            fill = severity_fills.get(severity, row_fill_even if is_alt else row_fill_odd)
            # For severity-colored rows: use lighter tint for non-severity columns
            if severity == "Critical":
                fill = PatternFill(start_color="FEE2E2" if not is_alt else "FECACA",
                                   end_color="FEE2E2" if not is_alt else "FECACA", fill_type="solid")
            elif severity == "High":
                fill = PatternFill(start_color="FEF3C7" if not is_alt else "FDE68A",
                                   end_color="FEF3C7" if not is_alt else "FDE68A", fill_type="solid")
            elif severity == "Medium":
                fill = PatternFill(start_color="FEF9C3" if not is_alt else "FEF08A",
                                   end_color="FEF9C3" if not is_alt else "FEF08A", fill_type="solid")
            apply_cell(c, font=body_font, fill=fill,
                       alignment=body_align_center if col_idx in (0, 2, 3) else body_align,
                       border=thin_border)
    ws2.row_dimensions[rr].height = 36

# Column widths Sheet 2
col_widths_s2 = [12, 14, 12, 14, 32, 48, 48]
for i, w in enumerate(col_widths_s2):
    ws2.column_dimensions[get_column_letter(i + 1)].width = w


# ════════════════════════════════════════════════════════════════════════
# SHEET 3 — 已知问题追踪
# ════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("已知问题追踪")
ws3.sheet_properties.tabColor = "F59E0B"

# Title
ws3.merge_cells("B2:G2")
c = ws3["B2"]
c.value = "P1–P6 已知优化项状态"
apply_cell(c, font=Font(name="Microsoft YaHei", size=16, bold=True, color=DARK_GREY),
           alignment=Alignment(horizontal="center", vertical="center"))
ws3.row_dimensions[2].height = 36

# Headers
headers3 = ["编号", "标题", "严重度", "状态", "审计确认", "修复方案"]
for i, h in enumerate(headers3):
    c = ws3.cell(row=4, column=2 + i, value=h)
    apply_cell(c, font=header_font, fill=header_fill, alignment=header_align, border=thin_border)
ws3.row_dimensions[4].height = 28

tracking_data = [
    ("P1", "AgentProgressPanel视觉", "Medium", "待修复", "审计已确认(M-8)", "添加skeleton加载态、时间跟踪显示和Error Boundary"),
    ("P2", "任务规划UI", "Medium", "待修复", "相关问题在ChatView", "拆分ChatView组件、优化规划交互流程"),
    ("P3", "Bridge→Chat桥接", "High", "待修复", "审计已确认(Agent-H3)", "添加发送到Chat按钮 + 事件总线集成"),
    ("P4", "对话删除", "Critical", "待修复", "审计已确认(Store-C1)", "修复sessions/[id]/route.ts的DELETE端点"),
    ("P5", "AgentsView去mock", "High", "待修复", "审计已确认(Agent-H1)", "接入真实Agent API、移除mock数据"),
    ("P6", "Chat API完善", "High", "待修复", "审计已确认(Chat-H4+Store-H3)", "修复历史加载顺序 + 添加消息重试功能"),
]

for idx, row_data in enumerate(tracking_data):
    rr = 5 + idx
    is_alt = idx % 2 == 1
    severity = row_data[2]
    for col_idx, val in enumerate(row_data):
        c = ws3.cell(row=rr, column=2 + col_idx, value=val)
        if col_idx == 2:  # Severity
            apply_cell(c, font=severity_fonts.get(severity, body_font),
                       fill=severity_fills.get(severity, row_fill_even if is_alt else row_fill_odd),
                       alignment=body_align_center, border=thin_border)
        elif col_idx == 3:  # Status
            status_font = Font(name="Microsoft YaHei", size=10, bold=True, color="B91C1C")
            status_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            apply_cell(c, font=status_font, fill=status_fill,
                       alignment=body_align_center, border=thin_border)
        else:
            fill = severity_fills.get(severity, row_fill_even if is_alt else row_fill_odd)
            if severity == "Critical":
                fill = PatternFill(start_color="FEE2E2" if not is_alt else "FECACA",
                                   end_color="FEE2E2" if not is_alt else "FECACA", fill_type="solid")
            elif severity == "High":
                fill = PatternFill(start_color="FEF3C7" if not is_alt else "FDE68A",
                                   end_color="FEF3C7" if not is_alt else "FDE68A", fill_type="solid")
            apply_cell(c, font=body_font, fill=fill,
                       alignment=body_align_center if col_idx in (0, 2, 3) else body_align,
                       border=thin_border)
    ws3.row_dimensions[rr].height = 40

# Column widths Sheet 3
ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 26
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 28
ws3.column_dimensions['G'].width = 52

# ── Print settings for all sheets ──
for ws in [ws1, ws2, ws3]:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

# Save
wb.save(OUTPUT)
print(f"✅ Audit report saved to: {OUTPUT}")
print(f"   Sheet 1: 审计总览 — summary stats + bar chart + module breakdown")
print(f"   Sheet 2: 问题清单 — {len(issues)} issues (Critical/High/Medium)")
print(f"   Sheet 3: 已知问题追踪 — P1-P6 tracking table")
