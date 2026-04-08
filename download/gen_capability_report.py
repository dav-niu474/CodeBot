import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============ Color palette ============
GREEN = "27AE60"
RED = "E74C3C"
AMBER = "F39C12"
BLUE = "2C3E50"
LIGHT_GREEN = "E8F5E9"
LIGHT_RED = "FFEBEE"
LIGHT_AMBER = "FFF8E1"
LIGHT_BLUE = "E3F2FD"
WHITE = "FFFFFF"
DARK = "333333"
GREY_BORDER = "D0D0D0"

thin_border = Border(
    left=Side(style='thin', color=GREY_BORDER),
    right=Side(style='thin', color=GREY_BORDER),
    top=Side(style='thin', color=GREY_BORDER),
    bottom=Side(style='thin', color=GREY_BORDER),
)

header_font = Font(name='Times New Roman', color=WHITE, bold=True, size=11)
header_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

title_font = Font(name='Times New Roman', size=18, bold=True, color="000000")
title_align = Alignment(horizontal='left', vertical='center')

subtitle_font = Font(name='Times New Roman', size=12, bold=True, color=BLUE)
body_font = Font(name='Times New Roman', size=10, color="000000")
body_align = Alignment(vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def status_fill(status):
    if status == "✅ 真实可用":
        return PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    elif status in ("⚠️ 部分/受限", "⚠️ 不可达"):
        return PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
    elif status in ("❌ 未实现", "❌ 装饰性", "❌ Stub"):
        return PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    return None

def status_font(status):
    if status == "✅ 真实可用":
        return Font(name='Times New Roman', size=10, color=GREEN, bold=True)
    elif status in ("⚠️ 部分/受限", "⚠️ 不可达"):
        return Font(name='Times New Roman', size=10, color=AMBER, bold=True)
    elif status in ("❌ 未实现", "❌ 装饰性", "❌ Stub"):
        return Font(name='Times New Roman', size=10, color=RED, bold=True)
    return body_font

def write_table(ws, start_row, headers, data, col_widths):
    # Headers
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=i+1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    # Data
    for r, row_data in enumerate(data):
        for c, val in enumerate(row_data):
            cell = ws.cell(row=start_row+1+r, column=c+2, value=val)
            cell.font = body_font
            cell.alignment = body_align if c > 0 else center_align
            cell.border = thin_border
            if c == 1 and status_fill(val):
                cell.fill = status_fill(val)
                cell.font = status_font(val)
    # Column widths
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+2)].width = w

# ============ Sheet 1: 总览 ============
ws1 = wb.active
ws1.title = "总览"
ws1.sheet_properties.tabColor = BLUE

ws1['B2'] = "CodeBot V4.1.0 — Agent 能力全景评估报告"
ws1['B2'].font = title_font
ws1['B2'].alignment = title_align
ws1.row_dimensions[2].height = 36

ws1['B4'] = "评估日期: 2026-04-08"
ws1['B4'].font = Font(name='Times New Roman', size=10, color="666666")

# Summary stats
ws1['B6'] = "能力模块概览"
ws1['B6'].font = subtitle_font

summary_data = [
    ["核心对话 (单Agent)", "✅ 真实可用", "流式API支持14个工具的完整agentic loop"],
    ["工具调用系统", "⚠️ 部分/受限", "14/44 工具有真实执行器，30个仅有Schema"],
    ["技能系统 (Skills)", "❌ 未实现", "无 /src/lib/skills/ 目录，仅有类型定义"],
    ["MCP 协议集成", "❌ 装饰性", "有UI和Schema，零实现代码"],
    ["计划清单 (Plan)", "✅ 真实可用", "/api/plan 可生成结构化计划，PlanPanel UI完整"],
    ["对话前自动计划", "❌ 未实现", "需用户手动触发，非自动"],
    ["多Agent — Coordinator", "✅ 真实可用", "LLM任务分解 + 并行Worker + 结果聚合"],
    ["多Agent — Swarm", "✅ 真实可用", "批量并行 + 消息总线 + 共识聚合"],
    ["多Agent — Teammate", "❌ 未实现", "引擎代码存在但无API路由，前端不可达"],
    ["子代理 (sub-agent)", "❌ Stub", "工具返回 'not yet available'"],
    ["自主模式 (agentic loop)", "✅ 真实可用", "最多10轮工具循环 + 上下文压缩恢复"],
    ["上下文压缩", "✅ 真实可用", "4种策略：snip/auto/responsive/micro"],
    ["记忆系统", "✅ 真实可用", "4层架构，对话前后自动存取"],
    ["安全审批流", "❌ 未实现", "UI组件存在但stream route未接入权限检查"],
    ["取消/中断", "⚠️ 部分/受限", "前端可中断fetch，后端无法取消已启动的worker"],
]

write_table(ws1, 8,
    ["模块", "状态", "说明"],
    summary_data,
    [22, 16, 60]
)

# Stats box
row = 8 + len(summary_data) + 2
ws1.cell(row=row, column=2, value="统计汇总").font = subtitle_font
stats = [
    ["✅ 真实可用", sum(1 for r in summary_data if r[1]=="✅ 真实可用"), "6/15 (40%)"],
    ["⚠️ 部分/受限", sum(1 for r in summary_data if "⚠️" in r[1]), "2/15 (13%)"],
    ["❌ 未实现/装饰/Stub", sum(1 for r in summary_data if r[1].startswith("❌")), "7/15 (47%)"],
]
write_table(ws1, row+1,
    ["状态", "数量", "占比"],
    stats,
    [22, 10, 15]
)

# ============ Sheet 2: 工具明细 ============
ws2 = wb.create_sheet("工具执行器明细")
ws2.sheet_properties.tabColor = GREEN

ws2['B2'] = "44 工具执行器 — 实现状态明细"
ws2['B2'].font = title_font
ws2['B2'].alignment = title_align
ws2.row_dimensions[2].height = 36

tools_data = [
    ["bash", "核心", "✅ 真实可用", "child_process.exec，沙箱环境变量，风险/阻断模式检测"],
    ["file-read", "核心", "✅ 真实可用", "fs.readFile + 偏移/限制分页，行号，截断保护"],
    ["file-write", "核心", "✅ 真实可用", "fs.writeFile + 自动创建父目录"],
    ["file-edit", "核心", "✅ 真实可用", "查找替换 + replaceAll + 预览验证"],
    ["glob", "核心", "✅ 真实可用", "自定义目录遍历，跳过node_modules/.git，200结果上限"],
    ["grep", "核心", "✅ 真实可用", "递归正则搜索，二进制跳过，50结果上限"],
    ["web-search", "核心", "✅ 真实可用", "z-ai-web-dev-sdk web_search"],
    ["web-fetch", "核心", "⚠️ 部分/受限", "原生fetch，15s超时，无重定向限制/无cookie"],
    ["todo-write", "核心", "⚠️ 部分/受限", "内存Map存储，服务器重启丢失"],
    ["send-message", "核心", "⚠️ 部分/受限", "返回JSON标记，stream route未解析"],
    ["ask-user", "核心", "⚠️ 部分/受限", "返回[ASK_USER]标记，stream route未处理"],
    ["brief", "核心", "⚠️ 部分/受限", "纯文本截断，非AI摘要"],
    ["agent", "核心", "❌ Stub", "返回 'not yet available in V3'"],
    ["notebook-edit", "核心", "❌ Stub", "返回 'not yet implemented'"],
    ["mcp", "懒加载", "❌ 未实现", "无MCP客户端/传输/协议代码"],
    ["list-mcp-resources", "懒加载", "❌ 未实现", "同上"],
    ["read-mcp-resource", "懒加载", "❌ 未实现", "同上"],
    ["mcp-auth", "懒加载", "❌ 未实现", "同上"],
    ["tool-search", "懒加载", "❌ 未实现", "无执行器"],
    ["lsp", "懒加载", "❌ 未实现", "无LSP客户端"],
    ["skill", "懒加载", "❌ 未实现", "无 /src/lib/skills/ 目录"],
    ["enter-plan-mode", "懒加载", "❌ 未实现", "无执行器"],
    ["exit-plan-mode", "懒加载", "❌ 未实现", "无执行器"],
    ["enter-worktree", "懒加载", "❌ 未实现", "无执行器"],
    ["exit-worktree", "懒加载", "❌ 未实现", "无执行器"],
    ["task-create", "懒加载", "❌ 未实现", "无执行器"],
    ["task-get", "懒加载", "❌ 未实现", "无执行器"],
    ["task-list", "懒加载", "❌ 未实现", "无执行器"],
    ["task-output", "懒加载", "❌ 未实现", "无执行器"],
    ["task-stop", "懒加载", "❌ 未实现", "无执行器"],
    ["task-update", "懒加载", "❌ 未实现", "无执行器"],
    ["team-create", "懒加载", "❌ 未实现", "无执行器"],
    ["team-delete", "懒加载", "❌ 未实现", "无执行器"],
    ["synthetic-output", "懒加载", "❌ 未实现", "无执行器"],
    ["config", "懒加载", "❌ 未实现", "无执行器"],
    ["remote-trigger", "懒加载", "❌ 未实现", "无执行器"],
    ["schedule-cron", "懒加载", "❌ 未实现", "无执行器"],
    ["powershell", "Flag", "❌ 未实现", "无执行器"],
    ["sleep", "Flag", "❌ 未实现", "无执行器"],
    ["repl", "Flag", "❌ 未实现", "无执行器"],
    ["voice", "Flag", "❌ 未实现", "无执行器"],
    ["dream-task", "Flag", "❌ 未实现", "无执行器"],
    ["magic-docs", "Flag", "❌ 未实现", "无执行器"],
]

write_table(ws2, 4,
    ["工具名", "类别", "状态", "详情"],
    tools_data,
    [20, 10, 16, 60]
)

# ============ Sheet 3: 多Agent模式 ============
ws3 = wb.create_sheet("多Agent系统")
ws3.sheet_properties.tabColor = "E67E22"

ws3['B2'] = "多Agent系统 — 三种模式实现状态"
ws3['B2'].font = title_font
ws3['B2'].alignment = title_align
ws3.row_dimensions[2].height = 36

agent_data = [
    ["Coordinator", "API路由", "✅ 真实可用", "/api/agents/coordinator — TransformStream SSE"],
    ["Coordinator", "任务分解", "✅ 真实可用", "LLM调用decomposeTask() → JSON子任务数组"],
    ["Coordinator", "并行执行", "✅ 真实可用", "Promise.allSettled 批次并行（≤3并发）"],
    ["Coordinator", "结果聚合", "✅ 真实可用", "LLM调用aggregateResults() → 综合摘要"],
    ["Coordinator", "DB持久化", "✅ 真实可用", "AgentSession + Session + Message 全生命周期"],
    ["Coordinator", "工具调用", "❌ 未实现", "Worker为纯文本LLM，无工具定义传入"],
    ["Coordinator", "取消机制", "❌ 未实现", "前端中断fetch后，后端继续执行"],
    ["Swarm", "API路由", "✅ 真实可用", "/api/agents/swarm — TransformStream SSE"],
    ["Swarm", "并行执行", "✅ 真实可用", "Promise.allSettled 批次并行（≤4并发）"],
    ["Swarm", "消息总线", "⚠️ 部分/受限", "AgentMessageBus存在但为批序单向，非真正P2P"],
    ["Swarm", "共识聚合", "✅ 真实可用", "LLM聚合 + confidence指标"],
    ["Swarm", "DB持久化", "✅ 真实可用", "完整AgentSession生命周期"],
    ["Swarm", "工具调用", "❌ 未实现", "Peer为纯文本LLM，无工具定义"],
    ["Swarm", "取消机制", "❌ 未实现", "同Coordinator"],
    ["Teammate", "API路由", "❌ 未实现", "无 /api/agents/teammate/route.ts"],
    ["Teammate", "前端调用", "❌ 不可达", "ChatView仅路由coordinator/swarm"],
    ["Teammate", "引擎代码", "⚠️ 不可达", "runTeammateMode()存在但从未被调用"],
]

write_table(ws3, 4,
    ["模式", "能力项", "状态", "说明"],
    agent_data,
    [14, 12, 16, 60]
)

# ============ Sheet 4: 关键差距与建议 ============
ws4 = wb.create_sheet("差距分析与建议")
ws4.sheet_properties.tabColor = RED

ws4['B2'] = "关键差距分析与实施建议"
ws4['B2'].font = title_font
ws4['B2'].alignment = title_align
ws4.row_dimensions[2].height = 36

gap_data = [
    ["P0", "MCP 协议集成", "❌ 装饰性", "极高", "需实现MCP Client + stdio/SSE传输 + 服务器发现 + 工具注册", "前端有完整UI（类别/图标/安全规则），接上后端即可用"],
    ["P0", "技能系统", "❌ 未实现", "高", "创建 /src/lib/skills/，实现技能加载器+执行器+模板系统", "SkillDef类型已定义，API路由存在"],
    ["P1", "Teammate模式路由", "❌ 不可达", "中", "创建 /api/agents/teammate/route.ts，在ChatView添加分支", "引擎代码已存在，只需接线"],
    ["P1", "Worker Agent工具调用", "❌ 未实现", "高", "在coordinator/swarm的chatCompletion()中传入工具定义+执行循环", "工具执行器已有，需在worker中复用"],
    ["P1", "对话前自动计划", "❌ 未实现", "中", "在agentic loop前加auto-plan判断，复杂任务自动调/api/plan", "PlanPanel UI已完整"],
    ["P1", "ask-user工具修复", "⚠️ 部分/受限", "中", "在stream route中检测[ASK_USER]标记，暂停loop等待用户输入", "标记格式已定义，只需解析处理"],
    ["P2", "send-message工具", "⚠️ 部分/受限", "低", "在stream route中检测send_message JSON并转发为SSE事件", "类似tool_call_result处理"],
    ["P2", "todo持久化", "⚠️ 部分/受限", "低", "将内存Map替换为DB存储（Prisma Memory表）", "Memory表已存在且有API"],
    ["P2", "Agent取消机制", "❌ 未实现", "中", "将AbortSignal传入coordinator/swarm引擎，在批次间检查", "前端AbortController已就绪"],
    ["P3", "子代理 (agent工具)", "❌ Stub", "高", "实现子agent生成+执行+结果回收，复用coordinator引擎", "架构复杂度高，建议下一版本"],
    ["P3", "安全审批流", "❌ 未实现", "中", "stream route读取SecurityView规则，高风工具暂停等待审批", "ToolApprovalDialog UI已存在"],
    ["P3", "真流式输出", "❌ 未实现", "低", "将chatCompletion()从非流式改为流式，逐token输出", "用户体验优化，非功能阻塞"],
]

write_table(ws4, 4,
    ["优先级", "差距项", "当前状态", "实施难度", "实施路径", "备注"],
    gap_data,
    [8, 20, 14, 10, 45, 35]
)

# ============ Save ============
output = "/home/z/my-project/download/CodeBot_V4.1.0_Agent_Capability_Assessment.xlsx"
wb.save(output)
print(f"Saved: {output}")
